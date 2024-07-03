from datetime import datetime, timedelta
import os
import sys
import pandas as pd
from openpyxl import Workbook, load_workbook
class Attendance:
    def extract_date(self, file_path):
        index = None
        month = None
        with open(file_path) as attendance_log:
            for line in attendance_log:
                line = line.replace("\t", " ").replace("\n", " ").strip()
                index, month = tuple(line.split()[0:2])
                break
        index, month = 101 if index[0] == "1" else 201, int(month.split("-")[1])

        return tuple([index, month])

    def process_attendance_log(self, current_index, file_path):
        filtered_data = []

        with open(file_path) as attendance_log:
            attendance_data = [line.replace("\t", " ").replace("\n", " ").strip() for line in attendance_log]

        filtered = [line for line in attendance_data if line.startswith(str(current_index))]

        for i in range(len(filtered)):
            current_timestamp = filtered[i][12:17]  # Extract timestamp (assuming timestamps are at positions 12:17)

            if i < len(filtered) - 1 and current_timestamp == filtered[i + 1][12:17]:
                # Check if the next element has the same timestamp
                if not filtered_data or filtered_data[-1][12:17] != current_timestamp:
                    filtered_data.append(filtered[i])
            elif not filtered_data or filtered_data[-1][12:17] != current_timestamp:
                # If filtered_data is empty or the timestamps don't match, add the current element to filtered_data
                filtered_data.append(filtered[i])

        return filtered_data

    def extract_data(self, current_index, data_list):
        result_dict = {}
        user_id = str(current_index)

        for item in data_list:
            parts = item.split()

            # Check if parts has at least three elements before trying to access its elements
            if len(parts) >= 3:
                user_id = parts[0]
                date = parts[1]
                time = int(parts[2].split(":")[0])

                if user_id not in result_dict:
                    result_dict[user_id] = {}

                if time <12:
                    result_dict[user_id][date] = {"checkIn": parts[2], "checkOut": ""}
                else:
                    if date not in result_dict[user_id]:
                        result_dict[user_id][date] = {"checkIn": "", "checkOut": parts[2]}
                    else:
                        result_dict[user_id][date]["checkOut"] = parts[2]
            else:
                if user_id not in result_dict:
                    result_dict[user_id] = {}
                result_dict[user_id][date] = {"checkIn": "", "checkOut": ""}
        return result_dict

    def fill_missing_dates(self, result_dict, date_ranges):
        for user_id, dates in date_ranges.items():
            if len(dates) >= 1:
                earliest_date, latest_date = dates[0], dates[-1]

                # Extract the year and month from the earliest_date
                year, month, _ = map(int, earliest_date.split('-'))

                # Get the first day of the month
                start_date = datetime(year, month, 1)

                # Get the last day of the month
                end_date = datetime(year, month % 12 + 1, 1) - timedelta(days=1)

                # Generate all dates for the given month
                all_dates = [str(start_date + timedelta(days=i)).split()[0] for i in range((end_date - start_date).days + 1)]

                for date in all_dates:
                    if date not in result_dict[user_id]:
                        result_dict[user_id][date] = {"checkIn": "", "checkOut": ""}
        return result_dict


    # ... (existing code remains unchanged)
# ... (existing code remains unchanged)

    def create_csv(self, index, result_dict, month_input):
        start_date = datetime(datetime.now().year, month_input, 1)
        end_date = datetime(datetime.now().year, month_input % 12 + 1, 1) - timedelta(
            days=1) if month_input < 12 else datetime(datetime.now().year + 1, 1, 1) - timedelta(days=1)
        output_file_path = "output.xlsx"

        if os.path.exists(output_file_path):
            workbook = load_workbook(output_file_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active

            # Writing the initial header row with merged cells for dates
            header_row = ['Index']
            for date in pd.date_range(start=start_date, end=end_date, freq='D'):
                header_row.append(date.strftime('%Y-%m-%d'))
                header_row.append('')  # Empty cell for merging
            sheet.append(header_row)

        # Get existing dates from the header row
        existing_dates = [cell.value for cell in sheet[1]]
        n = len(existing_dates)
        # print(existing_dates)
        # exit()
        # Writing the data
        for user_id, user_data in result_dict.items():
            data_row = [user_id]

            for date, times in user_data.items():
                if date in existing_dates:
                    # Find the column index for the existing date
                    col_index = existing_dates.index(date) + 1
                    data_row.extend([times['checkIn'], times['checkOut']])
                else:
                    # Add a new column for the new date in the header row
                    sheet.cell(row=1, column=len(existing_dates) + 1, value=date.strftime('%Y-%m-%d'))
                    sheet.cell(row=1, column=len(existing_dates) + 2, value='')  # Empty cell for merging
                    data_row.extend(['', ''])  # Add empty values for previous dates
                    data_row.extend(['', ''])  # Add empty values for the new date

            # Find the row index for the user_id
            row_index = len(sheet['A']) + 1
            # Write the data to the correct columns
            for col_index, value in enumerate(data_row, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)


        workbook.save(output_file_path)
        # print(f'Data appended to the XLSX file successfully.')


    def find_date_ranges(self, result_dict):
        return {user_id: sorted(result_dict[user_id].keys()) for user_id in result_dict}

    def execute_script(self):
        try:
            file_name = 'Admin' #input('Enter File Name: ')
            file_path = file_name + '.dat'
            current_index, month_input = self.extract_date(file_path)

            end_index = 199 if current_index == 101 else 300

            while current_index <= end_index:
                data_list = self.process_attendance_log(current_index, file_path)
                result_dict = self.extract_data(current_index, data_list)

                if str(current_index) in result_dict:
                    date_ranges = self.find_date_ranges(result_dict)
                    result_dict = self.fill_missing_dates(result_dict, date_ranges)
                    sorted_data = {k: dict(sorted(v.items(), key=lambda x: x[0])) for k, v in result_dict[str(current_index)].items()}
                    sorted_data = dict(sorted(sorted_data.items(), key=lambda x: x[0]))
                    s = current_index
                    data_dict = {}
                    data_dict[str(s)] = sorted_data
                    self.create_csv(current_index, data_dict, month_input)
                else:
                    print(f"No data found for index {current_index}")

                current_index += 1

            print("CSV file created successfully.")
        except Exception as e:
            print(e)

# ... (rest of the code remains unchanged)

   
    
if __name__ == "__main__":
    obj = Attendance()
    obj.execute_script()
    